import pandas as pd
import pathlib
import openpyxl
from math import ceil

path = pathlib.Path(__file__).parent.resolve().absolute()

dataframe = openpyxl.load_workbook(str(path)+'\Westmarch tables.xlsx', data_only=True)

dataframe1 = dataframe.active

# Read the number of tables and size
TableLocations = {}
for row in range(22-1, 22+2):
    #lineString = ""
    
    data = [None, None, None, None, None, None]
    
    for i, col in enumerate(dataframe1.iter_cols(19, 24)):
        data[i] = col[row].value
    if not data[0]: break
    tableName = data[0]
    startColumn = int(data[2])
    endColumn = int(data[3])
    startRow = int(data[4])
    endRow = int(data[5])
    TableLocations[tableName] = (startRow, endRow, startColumn, endColumn)
print(TableLocations)

def readTable(tableCoordinates):
    startRow = int(tableCoordinates[0])
    endRow = int(tableCoordinates[1])
    startColumn = int(tableCoordinates[2])
    endColumn = int(tableCoordinates[3])
    
    cells = [[None] * (endColumn - startColumn + 1) for i in range(startRow, endRow+1)]

    nameLengths = [None] * (endColumn - startColumn + 1)
    names = [None] * (endColumn - startColumn + 1)
    bothCount = 0
    for row in range(startRow-1, startRow):
        sameRows = 0
        for i, col in enumerate(dataframe1.iter_cols(startColumn, endColumn)):
            nameLengths[i] = max(len(str(col[row].value)) if col[row].value else 0, 6)
            names[i] = str(col[row].value) if col[row].value else ""
            if(bothCount > 1):
                cells[0][i-1] = names[i]
                bothCount = 1
            elif(bothCount > 0):
                cells[0][i-2] = cells[0][i-2] + " and " + names[i]
                nameLengths[i-2] = max(len(cells[0][i-2]), 6)
                bothCount = 0
            if(names[i] == 'Both =>'):
                bothCount = 2
            if(len(names[i]) == 0):
                sameRows += 1
            if(sameRows > 0 and (len(names[i]) > 0 or i == endColumn - startColumn)):
                nameLengths[i-sameRows] = max(ceil(nameLengths[i-1] / (sameRows+1)), 6)
                sameRows = 0
            cells[0][i] = names[i]
    
    lines = []
    for row in range(startRow, endRow):
        j = row - startRow + 1
        for i, col in enumerate(dataframe1.iter_cols(startColumn, endColumn)):
            format = str(col[row].number_format).replace("\\", "").replace("0", "")
            if(row == startRow):
                nameLengths[i] += len(format)
            cells[j][i] = str(ceil(int(col[row].value)))+format if not col[row].value == None else ""
            if(bothCount > 1):
                cells[j][i-1] = cells[j][i]
                bothCount = 1
            elif(bothCount > 0):
                cells[j][i-2] = cells[j][i-2] + " and " + cells[j][i]
                bothCount = 0
            if(names[i] == 'Both =>'):
                bothCount = 2

    for row in cells:
        lineString = ""
        for i, cell in enumerate(row):
            lineString += cell.rjust(max(nameLengths[i], len(cell)+2))
        lines.append(lineString)

    for line in lines:
        print(line)
for (name, value) in TableLocations.items():
    print("\n"+name)
    readTable(value)